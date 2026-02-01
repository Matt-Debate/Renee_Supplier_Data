#!/usr/bin/env python3
"""
Renee(A)->Renee(B) converter (hockey sticks only)

What it does
- Reads Spreadsheet A and extracts ONLY the 3 hockey-stick blocks:
    Block 1: B–F  (Model, Blade, Flex, Left, Right)
    Block 2: H–L
    Block 3: N–R
  Data expected to start row 5 with headers row 4.

- Handles merged cells in Model/Blade by using merged range top-left value,
  then fill-downs Model/Blade while scanning.

- Defect rule (default ON):
    If a Left/Right quantity cell contains any non-ASCII characters (often Chinese notes),
    that cell is excluded (treated as blank).

- Uses Spreadsheet B as a template and preserves formatting/styles.
  Adds a Style/Color column between Model and Blade if missing:
    B: Model
    C: Style/Color
    D: Blade
    E: Flex
    F: Left
    G: Right

- Extracts a trailing parenthetical suffix from Model (both ASCII and full-width Chinese parentheses):
    "FT8 Pro (RED)" -> Model="FT8 Pro", Style="RED"
    "FT6（red,black.blue,green）" -> Model="FT6", Style="red,black.blue,green"

- Prevents Style "bleed" across models when template has blank style rows:
    If a new model begins and style is blank on that row, current_style resets to None.

Output
- Writes a new .xlsx preserving B formatting.
- Optional CSV diff between original B and generated output.

Usage
  python transform.py --a "A.xlsx" --b "B.xlsx" --out "out.xlsx"
"""

import argparse
import math
import re
from collections import defaultdict
from copy import copy as ccopy
from typing import Any, Dict, Tuple, Optional, List

from openpyxl import load_workbook


# -----------------------------
# Parsing helpers
# -----------------------------

def has_non_ascii(s: str) -> bool:
    return any(ord(ch) > 127 for ch in s)

def norm_text(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def norm_model(v: Any) -> Optional[str]:
    return norm_text(v)

def norm_blade(v: Any) -> Optional[str]:
    s = norm_text(v)
    return s.upper() if s else None

def norm_style(v: Any) -> Optional[str]:
    s = norm_text(v)
    return s.upper() if s else None

def split_model_and_style(model_raw: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Extract trailing parentheses content, supporting both ASCII () and full-width （）.

    Examples:
      "FT8 Pro (RED)" -> ("FT8 Pro", "RED")
      "FT6（red,black.blue,green）" -> ("FT6", "red,black.blue,green")
      "Flylite USA Flag     （Tracer axis）" -> ("Flylite USA Flag", "Tracer axis")
    """
    s = norm_model(model_raw)
    if not s:
        return None, None

    # Normalize full-width parentheses to ASCII and trim
    s_norm = s.replace("（", "(").replace("）", ")").strip()

    # Only match a trailing (...) group
    m = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", s_norm)
    if not m:
        return s_norm, None

    base = (m.group(1) or "").strip() or s_norm
    style = (m.group(2) or "").strip() or None
    return base, style

def parse_flex(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(round(v))
    if isinstance(v, str):
        m = re.search(r"\d+", v)
        return int(m.group(0)) if m else None
    return None

def parse_qty(v: Any, defect_exclusion: bool = True) -> Optional[int]:
    """
    Quantity parsing:
    - numeric -> int
    - string with digits -> int(digits)
    - if defect_exclusion and string contains non-ascii -> excluded (None)
    """
    if v is None:
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        if math.isfinite(v) and abs(v - round(v)) < 1e-9:
            return int(round(v))
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        if defect_exclusion and has_non_ascii(s):
            return None
        m = re.search(r"\d+", s)
        return int(m.group(0)) if m else None
    return None


# -----------------------------
# Excel merged-cell helper
# -----------------------------

def merged_top_left_value(ws, cell) -> Any:
    """
    If cell is inside a merged range, return value of top-left cell of that range.
    Otherwise return cell.value.
    """
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(rng.min_row, rng.min_col).value
    return cell.value


# -----------------------------
# Ensure Style/Color column exists in B
# -----------------------------

def ensure_style_column(ws) -> None:
    """
    Ensures there is a Style/Color column inserted between Model and Blade.

    Original template:
      B: Model, C: Blade, D: Flex, E: Left, F: Right

    After:
      B: Model, C: Style/Color, D: Blade, E: Flex, F: Left, G: Right
    """
    header_c = ws.cell(1, 3).value
    header_c_s = str(header_c).strip().lower() if header_c is not None else ""

    # If column C already looks like style/color, do nothing
    if "style" in header_c_s or "color" in header_c_s:
        return

    # Preserve old column C width if set
    old_c_dim = ws.column_dimensions.get("C", None)
    old_c_width = old_c_dim.width if old_c_dim else None

    # Insert new column C
    ws.insert_cols(3)
    ws.cell(1, 3).value = "Style/Color"

    # Copy formatting from column B into the new column C
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        src = ws.cell(r, 2)  # B
        dst = ws.cell(r, 3)  # new C
        dst.font = ccopy(src.font)
        dst.fill = ccopy(src.fill)
        dst.border = ccopy(src.border)
        dst.alignment = ccopy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = ccopy(src.protection)

    if old_c_width is not None:
        ws.column_dimensions["C"].width = old_c_width


# -----------------------------
# Extract inventory from A
# -----------------------------

def build_inventory_from_a(
    path_a: str,
    defect_exclusion: bool = True
) -> Dict[Tuple[str, Optional[str], str, int], Tuple[Optional[int], Optional[int]]]:
    wb_a = load_workbook(path_a, data_only=True)
    ws_a = wb_a.active

    blocks = [
        {"model_col": 2,  "blade_col": 3,  "flex_col": 4,  "left_col": 5,  "right_col": 6},    # B-F
        {"model_col": 8,  "blade_col": 9,  "flex_col": 10, "left_col": 11, "right_col": 12},   # H-L
        {"model_col": 14, "blade_col": 15, "flex_col": 16, "left_col": 17, "right_col": 18},   # N-R
    ]

    summed = defaultdict(lambda: {"L": 0, "R": 0})

    for blk in blocks:
        current_model_raw = None
        current_blade_raw = None

        for r in range(5, ws_a.max_row + 1):
            model_v = merged_top_left_value(ws_a, ws_a.cell(r, blk["model_col"]))
            blade_v = merged_top_left_value(ws_a, ws_a.cell(r, blk["blade_col"]))
            flex_v  = ws_a.cell(r, blk["flex_col"]).value
            left_v  = ws_a.cell(r, blk["left_col"]).value
            right_v = ws_a.cell(r, blk["right_col"]).value

            if norm_model(model_v):
                current_model_raw = model_v
            if norm_blade(blade_v):
                current_blade_raw = blade_v

            model_use = model_v if norm_model(model_v) else current_model_raw
            blade_use = blade_v if norm_blade(blade_v) else current_blade_raw

            model_base, style = split_model_and_style(model_use)
            style = norm_style(style)
            blade = norm_blade(blade_use)
            flex = parse_flex(flex_v)

            if model_base is None or blade is None or flex is None:
                continue

            L = parse_qty(left_v, defect_exclusion=defect_exclusion)
            R = parse_qty(right_v, defect_exclusion=defect_exclusion)

            key = (model_base, style, blade, flex)
            if L is not None:
                summed[key]["L"] += L
            if R is not None:
                summed[key]["R"] += R

    inv: Dict[Tuple[str, Optional[str], str, int], Tuple[Optional[int], Optional[int]]] = {}
    for k, v in summed.items():
        L = v["L"] if v["L"] > 0 else None
        R = v["R"] if v["R"] > 0 else None
        inv[k] = (L, R)

    return inv


# -----------------------------
# Apply inventory to B template
# -----------------------------

def apply_to_b_template(
    path_b: str,
    out_path: str,
    inv: Dict[Tuple[str, Optional[str], str, int], Tuple[Optional[int], Optional[int]]],
    filldown_b_template: bool = True
) -> None:
    wb_out = load_workbook(path_b)  # preserve styles
    ws = wb_out.active

    ensure_style_column(ws)

    # After ensure_style_column, columns are:
    # B Model, C Style, D Blade, E Flex, F Left, G Right
    COL_MODEL = 2
    COL_STYLE = 3
    COL_BLADE = 4
    COL_FLEX  = 5
    COL_LEFT  = 6
    COL_RIGHT = 7

    def row_has_any(r: int) -> bool:
        for c in (COL_MODEL, COL_STYLE, COL_BLADE, COL_FLEX, COL_LEFT, COL_RIGHT):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                return True
        return False

    last_row = ws.max_row
    while last_row > 1 and not row_has_any(last_row):
        last_row -= 1

    # Clear Left/Right to avoid template values leaking
    for r in range(2, last_row + 1):
        ws.cell(r, COL_LEFT).value = None
        ws.cell(r, COL_RIGHT).value = None

    current_model: Optional[str] = None
    current_style: Optional[str] = None
    current_blade: Optional[str] = None

    for r in range(2, last_row + 1):
        model_cell = ws.cell(r, COL_MODEL)
        style_cell = ws.cell(r, COL_STYLE)
        blade_cell = ws.cell(r, COL_BLADE)
        flex_cell  = ws.cell(r, COL_FLEX)
        left_cell  = ws.cell(r, COL_LEFT)
        right_cell = ws.cell(r, COL_RIGHT)

        # Read model/style, allowing older templates that had "Model (STYLE)" in column B
        base_model_here, style_from_model = split_model_and_style(model_cell.value)
        style_here = norm_style(style_cell.value) or norm_style(style_from_model)
        blade_here = norm_blade(blade_cell.value)
        flex = parse_flex(flex_cell.value)

        # If we extracted trailing parentheses from Model cell, rewrite it to base model
        if base_model_here and style_from_model:
            model_cell.value = base_model_here

        # IMPORTANT: stop style bleeding across models
        model_changed = base_model_here is not None and base_model_here != current_model
        if model_changed and not style_here:
            current_style = None

        # Update current tracking
        if base_model_here:
            current_model = base_model_here
        if style_here:
            current_style = style_here
        if blade_here:
            current_blade = blade_here

        # Fill down within template if requested
        model = base_model_here or (current_model if filldown_b_template else None)
        style = style_here or (current_style if filldown_b_template else None)
        blade = blade_here or (current_blade if filldown_b_template else None)

        if filldown_b_template:
            if (model_cell.value is None or str(model_cell.value).strip() == "") and model:
                model_cell.value = model
            if (style_cell.value is None or str(style_cell.value).strip() == "") and style:
                style_cell.value = style
            if (blade_cell.value is None or str(blade_cell.value).strip() == "") and blade:
                blade_cell.value = blade

        if model is None or blade is None or flex is None:
            continue

        key = (model, style, blade, flex)
        newL, newR = inv.get(key, (None, None))

        left_cell.value = newL
        right_cell.value = newR

    wb_out.save(out_path)


# -----------------------------
# Diff report (optional)
# -----------------------------

def write_diff_csv(path_b_original: str, path_b_generated: str, diff_csv: str) -> None:
    import csv

    wb_o = load_workbook(path_b_original, data_only=True)
    wb_g = load_workbook(path_b_generated, data_only=True)
    ws_o = wb_o.active
    ws_g = wb_g.active

    # Normalize structure so comparisons line up
    ensure_style_column(ws_o)
    ensure_style_column(ws_g)

    cols = [
        (2, "Model"),
        (3, "Style/Color"),
        (4, "Blade"),
        (5, "Flex"),
        (6, "Left"),
        (7, "Right"),
    ]

    def row_has_any(ws, r: int) -> bool:
        return any(ws.cell(r, c).value not in (None, "") for c, _ in cols)

    last_row = ws_o.max_row
    while last_row > 1 and not row_has_any(ws_o, last_row):
        last_row -= 1

    diffs: List[Dict[str, Any]] = []
    for r in range(1, last_row + 1):
        for c, name in cols:
            o = ws_o.cell(r, c).value
            g = ws_g.cell(r, c).value

            o_blank = (o is None) or (isinstance(o, str) and o.strip() == "")
            g_blank = (g is None) or (isinstance(g, str) and g.strip() == "")
            if o_blank and g_blank:
                continue

            if o != g:
                diffs.append({"row": r, "col": name, "orig_B": o, "generated": g})

    with open(diff_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["row", "col", "orig_B", "generated"])
        w.writeheader()
        w.writerows(diffs)


# -----------------------------
# Public API for Streamlit
# -----------------------------

def transform_files(
    path_a: str,
    path_b: str,
    out_path: str,
    defect_exclusion: bool = True,
    filldown_b_template: bool = True,
    diff_csv: Optional[str] = None
) -> None:
    inv = build_inventory_from_a(path_a, defect_exclusion=defect_exclusion)
    apply_to_b_template(path_b, out_path, inv, filldown_b_template=filldown_b_template)
    if diff_csv:
        write_diff_csv(path_b, out_path, diff_csv)


# -----------------------------
# CLI
# -----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--a", required=True, help="Path to Spreadsheet A (.xlsx)")
    ap.add_argument("--b", required=True, help="Path to Spreadsheet B template (.xlsx)")
    ap.add_argument("--out", required=True, help="Output path for generated .xlsx")
    ap.add_argument("--no-defect-exclusion", action="store_true",
                    help="If set, quantities with Chinese/non-ascii annotations will NOT be excluded.")
    ap.add_argument("--no-filldown-b-template", action="store_true",
                    help="If set, do NOT fill down blank cells inside B template.")
    ap.add_argument("--diff-csv", default=None,
                    help="Optional path to write a CSV diff report comparing provided B vs generated output.")

    args = ap.parse_args()

    transform_files(
        path_a=args.a,
        path_b=args.b,
        out_path=args.out,
        defect_exclusion=(not args.no_defect_exclusion),
        filldown_b_template=(not args.no_filldown_b_template),
        diff_csv=args.diff_csv,
    )

if __name__ == "__main__":
    main()